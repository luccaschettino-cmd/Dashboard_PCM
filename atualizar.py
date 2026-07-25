"""
=======================================================
  PCM Produtos Viçosa — Gerador de dados para o site
  Autor: Lucca Schettino · PCM
  Uso:   python atualizar.py
=======================================================
"""

import json, sys, os, re
from datetime import datetime, timedelta
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("❌  Instale com:  pip install openpyxl")
    sys.exit(1)

PLANILHA   = "PCM_OS_Intranet.xlsx"
ABA_DADOS  = "Dados Brutos"
SAIDA_JSON = "dados.json"
PRAZO_DIAS = 30

MESES_PT = {
    1:"Janeiro", 2:"Fevereiro", 3:"Março",    4:"Abril",
    5:"Maio",    6:"Junho",     7:"Julho",     8:"Agosto",
    9:"Setembro",10:"Outubro", 11:"Novembro", 12:"Dezembro"
}

def limpar(v):
    return "" if v is None else str(v).strip()

def para_data(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if hasattr(v, "year"): return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y","%Y-%m-%d","%d-%m-%Y","%m/%d/%Y"):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    return None

def para_tempo(v):
    """Aceita: 2, 2.0, '2,0d', '2.0d', '137,0d' etc."""
    if v is None: return None
    s = str(v).replace(",", ".").replace("d","").replace("D","").strip()
    try: return float(s)
    except: return None

def mes_de(v):
    """Aceita número (1-12) ou nome ('Janeiro') ou abreviação ('Jan')"""
    if v is None: return None, ""
    # se for número
    try:
        n = int(float(str(v)))
        if 1 <= n <= 12:
            return n, MESES_PT[n]
    except: pass
    # se for texto
    s = str(v).strip().capitalize()
    for num, nome in MESES_PT.items():
        if nome.startswith(s[:3]):
            return num, nome
    return None, ""

def ler_planilha(caminho):
    print(f"\n📂  Abrindo: {caminho}")
    if not os.path.exists(caminho):
        print(f"❌  Arquivo não encontrado: {caminho}")
        sys.exit(1)
    try:
        wb = openpyxl.load_workbook(caminho, data_only=True)
    except Exception as e:
        print(f"❌  Erro ao abrir: {e}"); sys.exit(1)

    if ABA_DADOS not in wb.sheetnames:
        print(f"❌  Aba '{ABA_DADOS}' não encontrada.")
        print(f"    Abas: {wb.sheetnames}"); sys.exit(1)

    ws = wb[ABA_DADOS]

    # descobre linha dos cabeçalhos (procura linha com "Código" ou "Situação")
    header_row = None
    for r in range(1, 10):
        vals = [limpar(ws.cell(row=r, column=c).value).lower()
                for c in range(1, ws.max_column+1)]
        if any("código" in v or "codigo" in v or "situação" in v or "situacao" in v
               for v in vals):
            header_row = r
            break

    if not header_row:
        print("❌  Não encontrei a linha de cabeçalhos.")
        sys.exit(1)

    headers = [limpar(ws.cell(row=header_row, column=j).value).lower()
               for j in range(1, ws.max_column+1)]

    def col(*nomes):
        for nome in nomes:
            for i, h in enumerate(headers):
                if nome in h:
                    return i
        return None

    idx = {
        "codigo":     col("código","codigo"),
        "situacao":   col("situação","situacao"),
        "prioridade": col("prioridade"),
        "abertura":   col("abertura"),
        "conclusao":  col("conclusão","conclusao"),
        "contagem":   col("contagem"),
        "tempo":      col("tempo"),
        "mes":        col("mês de","mes de","mês","mes"),
    }
    print(f"    Cabeçalhos na linha {header_row}: {[h for h in headers if h]}")

    ordens = []
    validas = 0

    for row in ws.iter_rows(min_row=header_row+1, values_only=True):
        if all(v is None or str(v).strip()=='' for v in row):
            continue

        def get(chave):
            i = idx.get(chave)
            return row[i] if i is not None and i < len(row) else None

        codigo     = limpar(get("codigo"))
        situacao   = limpar(get("situacao"))
        prioridade = limpar(get("prioridade"))
        dt_aber    = para_data(get("abertura"))
        dt_conc    = para_data(get("conclusao"))
        tempo      = para_tempo(get("tempo"))
        mes_num, mes_nome = mes_de(get("mes"))

        # se não veio mês, tenta extrair da data de abertura
        if not mes_num and dt_aber:
            mes_num  = dt_aber.month
            mes_nome = MESES_PT[mes_num]

        # "Entra na contagem?" — se vier "Não"/"Nao"/"N" a OS é ignorada completamente
        contagem_raw = limpar(get("contagem")).lower()
        if contagem_raw.startswith(("não","nao","n")):
            continue   # pula essa OS, como se não existisse

        # calcula tempo se não veio
        if tempo is None and dt_aber and dt_conc:
            tempo = (dt_conc - dt_aber).days

        # concluída no prazo
        concluida = situacao.lower().startswith(("conclu","finaliz"))
        concluida_prazo = (
            concluida and
            tempo is not None and tempo <= PRAZO_DIAS
        )

        if not codigo and not dt_aber:
            continue

        validas += 1
        ordens.append({
            "codigo":          codigo,
            "situacao":        situacao,
            "prioridade":      prioridade,
            "dt_abertura":     dt_aber.isoformat() if dt_aber else None,
            "dt_conclusao":    dt_conc.isoformat() if dt_conc else None,
            "entra_contagem":  True,
            "tempo_dias":      tempo,
            "concluida_prazo": concluida_prazo,
            "mes_num":         mes_num,
            "mes_nome":        mes_nome,
        })

    print(f"✅  Ordens válidas: {validas}")
    return ordens

def consolidar(ordens):
    meses = defaultdict(lambda:{
        "abertas":0,"conc30":0,"nconc30":0,
        "tempo_soma":0.0,"tempo_n":0,
        "urgente":0,"alta":0,"normal":0
    })
    total = {"abertas":0,"conc30":0,"nconc30":0,
             "urgente":0,"alta":0,"normal":0,
             "tempo_soma":0.0,"tempo_n":0}

    for o in ordens:
        mn = o["mes_num"]
        if not mn: continue
        m = meses[mn]
        m["abertas"] += 1; total["abertas"] += 1

        prio = o["prioridade"].lower()
        if "urgente" in prio:
            m["urgente"]+=1; total["urgente"]+=1
        elif "alta" in prio:
            m["alta"]+=1; total["alta"]+=1
        else:
            m["normal"]+=1; total["normal"]+=1

        if o["concluida_prazo"]:
            m["conc30"]+=1; total["conc30"]+=1
            if o["tempo_dias"] is not None:
                m["tempo_soma"]+=o["tempo_dias"]; m["tempo_n"]+=1
                total["tempo_soma"]+=o["tempo_dias"]; total["tempo_n"]+=1
        elif not o["concluida_prazo"]:
            m["nconc30"]+=1; total["nconc30"]+=1

    resumo = []
    for mn in sorted(meses.keys()):
        m = meses[mn]
        ind   = round(m["conc30"]/m["abertas"]*100,2) if m["abertas"] else 0
        tmedio= round(m["tempo_soma"]/m["tempo_n"],2) if m["tempo_n"] else 0
        resumo.append({
            "mes_num":    mn, "mes": MESES_PT[mn],
            "abertas":    m["abertas"], "conc30":  m["conc30"],
            "nconc30":    m["nconc30"], "tempo_medio": tmedio,
            "indicador":  ind, "urgente": m["urgente"],
            "alta":       m["alta"],    "normal":  m["normal"],
        })

    ind_t = round(total["conc30"]/total["abertas"]*100,2) if total["abertas"] else 0
    tm_t  = round(total["tempo_soma"]/total["tempo_n"],2) if total["tempo_n"] else 0
    totais = {
        "abertas":total["abertas"],"conc30":total["conc30"],
        "nconc30":total["nconc30"],"tempo_medio":tm_t,
        "indicador":ind_t,"urgente":total["urgente"],
        "alta":total["alta"],"normal":total["normal"],
    }
    return resumo, totais

def main():
    print("="*52)
    print("  PCM Produtos Viçosa — Gerador de dados do site")
    print("="*52)

    caminho = sys.argv[1] if len(sys.argv) > 1 else PLANILHA
    ordens  = ler_planilha(caminho)

    if not ordens:
        print("⚠️  Nenhuma ordem encontrada.")
        sys.exit(1)

    resumo, totais = consolidar(ordens)

    hoje = datetime.today().date()
    os_15d = sum(1 for o in ordens
        if o["dt_abertura"] and
        datetime.fromisoformat(o["dt_abertura"]).date() >= hoje - timedelta(days=15))

    payload = {
        "_gerado_em": datetime.now().isoformat(timespec="seconds"),
        "_fonte":     caminho,
        "intranet": {
            "resumo_mensal":  resumo,
            "totais":         totais,
            "os_ultimos_15d": os_15d,
            "meta_indicador": 90,
        }
    }

    with open(SAIDA_JSON,"w",encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"\n✅  JSON gerado: {SAIDA_JSON}")
    print(f"    Meses processados  : {len(resumo)}")
    print(f"    Total de ordens    : {totais['abertas']}")
    print(f"    Concluídas em 30d  : {totais['conc30']}")
    print(f"    Indicador geral    : {totais['indicador']}%")
    print(f"    Urgente/Alta/Normal: {totais['urgente']}/{totais['alta']}/{totais['normal']}")
    print(f"\n📤  Suba o '{SAIDA_JSON}' no GitHub (mesma pasta do index.html)")
    print("="*52)

if __name__ == "__main__":
    main()
